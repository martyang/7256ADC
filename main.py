
import os
import time
import serial
import serial.tools.list_ports_windows
from openpyxl.reader.excel import load_workbook
from serial import SerialException
from openpyxl import Workbook
from chamber import Chamber


def getWorkBook():
    if not os.path.exists('./result.xlsx'):
        wkbook = Workbook()
        wksheet = wkbook.active
        wksheet.merge_cells('C1:D1')
        wksheet.merge_cells('E1:F1')
        wksheet.merge_cells('G1:H1')
        wksheet.merge_cells('I1:J1')
        wksheet['C1'] = "#1"
        wksheet['E1'] = "#2"
        wksheet['G1'] = "#3"
        wksheet['I1'] = "#4"
        wksheet.append(
            ['温箱温度', '传感器温度', 'voltage', 'temperature', 'voltage', 'temperature', 'voltage', 'temperature',
             'voltage', 'temperature'])
        wkbook.save('result.xlsx')

    wkbook = load_workbook('result.xlsx')
    return wkbook


def serialData2Excel(result_list):
    """

    :return:
    """
    portlist = serial.tools.list_ports_windows.comports()
    result_list = result_list
    for item in portlist:
        if 'com1' not in item:
            comport = str(item).split('-')[0]
            print('当前测试端口:' + comport)
            voltage = 'None'
            temp = 'None'
            try:
                myserial = serial.Serial(comport, 115200, timeout=5)
                time.sleep(3)
                while myserial.inWaiting():
                    recv_data = myserial.readline().decode('utf-8')
                    print(recv_data)
                    if 'ntc_drv_read_temp voltage value' in recv_data:
                        voltage = recv_data.split(':')[3].split(' ')[1]
                        result_list.append(int(voltage))
                    elif 'temperature' in recv_data:
                        temp = recv_data.split(':')[3].split(' ')[1]
                        result_list.append(float(temp))
                    if 'None' != voltage and 'None' != temp:
                        break
                myserial.close()
            except SerialException:
                print('无法打开串口')
            except UnicodeDecodeError:
                print('decode error')
    workbook = getWorkBook()
    workbook.active.append(result_list)
    workbook.save('result.xlsx')


def AdcTempTest():
    config_file = open(os.getcwd() + "\\config.txt", 'rb')
    content = config_file.read().decode('utf-8')
    start_temp = int(content.split('\n')[0].strip().split(':')[1])
    end_temp = int(content.split('\n')[1].strip().split(':')[1])
    step = int(content.split('\n')[2].strip().split(':')[1])
    addr = content.split('\n')[3].strip().split(';')[1]
    print(f'Test info {start_temp},{end_temp},{step}')
    print('Test start')

    chamber = Chamber(addr)
    chamber.powerOn()
    for t in range(start_temp, end_temp+1, step):
        chamber.setTemp(t)
        time.sleep(60)
        result_list = [t, '']
        while chamber.getAvgTof10S() != t:
            cuT = chamber.getCurrentT()
            print('Chamber current temp %f ℃' % cuT)
            time.sleep(60)
        else:
            cuT = chamber.getCurrentT()
            print('Testing  temp %f ℃' % cuT)
            # 温度达到要求，开始读取数据
            serialData2Excel(result_list)
            time.sleep(1)
    print("Test completed")
    chamber.powerOff()


if __name__ == '__main__':
    AdcTempTest()


